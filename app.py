import requests
import openpyxl
import time
wb = openpyxl.load_workbook("test1.xlsx")
sheet_obj = wb.active
while True:
	rows = sheet_obj.max_row
	for i in range(2,rows+1):
		cell_obj = sheet_obj.cell(row = i ,column = 1)
		temp = sheet_obj.cell(row = i,column = 2)
		update = sheet_obj.cell(row = i,column = 3)
		if update.value == 1:
			if temp.value == 'F':
				unit == "imperial"
			else:
				unit == "metric"
			web_url = requests.get("http://api.openweathermap.org/data/2.5/weather?appid=b13cff6465438dd33fd22463de9fbf34&mode=json&units={}&q={}".format(unit,cell_obj.value))
            sheet_obj.cell(column =4 , row = i,value = web_url.json()["main"]["temp"])
            sheet_obj.cell(column = 5,row = i,value = web_url.json()["main"]["humidity"])
            sheet_obj.cell(column = 6,row = i,value = web_url.json()["main"]["pressure"])
            sheet_obj.cell(column = 7,row = i,value = web_url.json()["weather"][0]["discription"])
        else:
        	continue
    wb.save("test1.xlsx")
    print("Information updated ")
    time.sleep(3)#sheet updated every 3 seconds
